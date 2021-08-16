from openpyxl import load_workbook
import re
import nltk
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
from collections import Counter
import pandas as pd
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from collections import Counter
import matplotlib.pyplot as plt 

wb = load_workbook(filename='Dataset3.xlsx')
ws = wb.active

tweetList = []
for row in ws["A"]:
    tweetList.append(row.value)

def cleanTweetForKeywords(text):
    text = re.sub(r'&amp;', '&', text)
    text = re.sub(r'\n', '', text)
    text = re.sub(r'@[A-Za-z0-9]+', '', text)
    text = re.sub(r'#', '', text)
    text = re.sub(r'https?:\/\/\S+', '', text)
    text = re.sub(r'[^\w\s]', '', text)
    text = text.lower()
    return text

# def cleanTweetForSentiment(text):
#     text = re.sub(r'&amp;', '&', text)
#     text = re.sub(r'@[A-Za-z0-9]+', '', text)
#     text = re.sub(r'#', '', text)
#     text = re.sub(r'https?:\/\/\S+', '', text)
#     return text

cleanedForKeywords = []
for tweet in tweetList:
    cleanedForKeywords.append(cleanTweetForKeywords(tweet))

def preprocess(text):
    tokenized = word_tokenize(text)
    stoppedWords = []
    for word in tokenized:
        if word not in stopwords.words('english'):
            stoppedWords.append(word)
    lemmatizer = WordNetLemmatizer()
    lemmatized = []
    for word in stoppedWords:
        lemmatized.append(lemmatizer.lemmatize(word))
    doubleLemmatized = []
    for word in lemmatized:
        doubleLemmatized.append(lemmatizer.lemmatize(word))
    return doubleLemmatized

# tokenized = [word_tokenize(sentence) for sentence in cleanedForKeywords]

# stoppedWords = []
# for sentence in tokenized:
#     newSentence = []
#     for word in sentence:
#         if word not in stopwords.words('english'):
#             newSentence.append(word)
#     stoppedWords.append(newSentence)

# lemmatizer = WordNetLemmatizer()
# lemmatized = []
# for sentence in stoppedWords:
#     newSentence = []
#     for word in sentence:
#         newSentence.append(lemmatizer.lemmatize(word))
#     lemmatized.append(newSentence)
# doubleLemmatized = []
# for sentence in lemmatized:
#     newSentence = []
#     for word in sentence:
#         newSentence.append(lemmatizer.lemmatize(word))
#     doubleLemmatized.append(newSentence)

# counterList = []
# for sentence in doubleLemmatized:
#     for word in sentence:
#         counterList.append(word)
    
# print(Counter(counterList))

# posTagged = []
# for sentence in doubleLemmatized:
#     postag = nltk.pos_tag(sentence)
#     for tup in postag:
#         posTagged.append(tup)

# df = pd.DataFrame(set(posTagged), columns=['word', 'pos'])
# df.to_excel("postags.xlsx", index=False)
# availablePos = set(df['pos'])
# print(availablePos)

def sentiment(text):
    analyzer = SentimentIntensityAnalyzer()
    sentiment_dict = analyzer.polarity_scores(text)
    return sentiment_dict

def sentimentText(sentimentDict):
    if sentimentDict >= 0.05:
        return 'positive'
    elif sentimentDict <= -0.05:
        return 'negative'
    else:
        return 'neutral'

compoundList = []
sentimentTextList = []
for data in cleanedForKeywords:
    compound = sentiment(data)['compound']
    compoundList.append(compound)
    sentimentTextList.append(sentimentText(compound))

df = pd.DataFrame(list(zip(cleanedForKeywords, compoundList, sentimentTextList)), columns=['Tweet', 'Compound', 'Sentiment'])
# df.to_excel('finalsentiment.xlsx', index=False)

positiveKeywords = []
negativeKeywords = []
neutralKeywords = []
for index in df.index:
    if df['Sentiment'][index] == 'positive':
        positiveKeywords.append(preprocess(df['Tweet'][index]))
    elif df['Sentiment'][index] == 'negative':
        negativeKeywords.append(preprocess(df['Tweet'][index]))
    elif df['Sentiment'][index] == 'neutral':
        neutralKeywords.append(preprocess(df['Tweet'][index]))

negativePosTagged = []
for sentence in negativeKeywords:
    postag = nltk.pos_tag(sentence)
    for tup in postag:
        negativePosTagged.append(tup)
df1 = pd.DataFrame(set(negativePosTagged), columns=['word', 'pos'])
df1.to_excel("negativepostags.xlsx", index=False)

positivePosTagged = []
for sentence in positiveKeywords:
    postag = nltk.pos_tag(sentence)
    for tup in postag:
        positivePosTagged.append(tup)
df2 = pd.DataFrame(set(positivePosTagged), columns=['word', 'pos'])
df2.to_excel("positivepostags.xlsx", index=False) 

neutralPosTagged = []
for sentence in neutralKeywords:
    postag = nltk.pos_tag(sentence)
    for tup in postag:
        neutralPosTagged.append(tup)
df3 = pd.DataFrame(set(neutralPosTagged), columns=['word', 'pos'])
df3.to_excel("neutralpostags.xlsx", index=False)

positives = sentimentTextList.count('positive')
negatives = sentimentTextList.count('negative')
neutrals = sentimentTextList.count('neutral')
fig = plt.figure()
plt.bar(['Positive', 'Negative', 'Neutral'], [positives, negatives, neutrals])
plt.xlabel("People's sentiment towards the tweet")
plt.ylabel("Number of tweets")
plt.show()

sustainableKeywords = ['sustainable', 'sustainability', 'sustainably']
mostOccured = ['sustainable', 'fastfashion', 'waste', 'child', 'fair', 'environment', 'greenwashing', 'union', 'safety', 'recycling', 'supplychain', 'slavery', 'labour', 'plastic', 'human', 'worker', 'ethical', 'organic', 'wage', 'diversity', 'transparency', 'woman', 'vegan', 'charity', 'conscious', 'boycott', 'green']
mostOccuredCompound = []
mostOccuredSentiment = []
for tweet in tweetList:
    tweet = tweet.lower()
    i = 0
    for word in mostOccured:
        if (word in tweet) and (i == 0):
            compound = sentiment(tweet)['compound']
            mostOccuredCompound.append(compound)
            mostOccuredSentiment.append(sentimentText(compound))
            i =+ 1

mostOccuredPositives = mostOccuredSentiment.count('positive')
mostOccuredNegatives = mostOccuredSentiment.count('negative')
mostOccuredNeutrals = mostOccuredSentiment.count('neutral')
fig = plt.figure()
plt.bar(['Positive', 'Negative', 'Neutral'], [mostOccuredPositives, mostOccuredNegatives, mostOccuredNeutrals])
plt.xlabel("Number of tweets")
plt.ylabel("Sustainability-related keywords")
plt.show()

# compoundList = []
# sentimentTextList = []

# keywordSentence = "landfill plastic disposable transparent transparency bangladesh fabric produce production producing bioplastic dump reusable pandemic india overtime fauxfur fur conscious recycle recycled recyclable hm slavery palestine israel greenwashing greenwash bio bioalternative child ecological plantbased alternative sustainable sustainability fast fastfashion climate illegal beijing unsustainable eco boycott labor labour wasteful waste ethical female exploitation vegan environment environmental immigrant worldearthday2021 zeroplast vintagesecondhand pakistan racist racism wage circular plasticpollution beatinghm overproduction discrimination blockchain activism disposal waste humanity itsapark cambodia overwork diversity safety representation labour consumption green organiccotton scandal climate fastfashion authenticity overproducing greenconsumerism maisiewilliams uneducated whomademyclothes supplychain supply supplier feminist forced neglected sourced repurposing organic upcycling donating authenticity overtime sustainably mistreating ecofriendly"
# keywords = word_tokenize(keywordSentence)

# print(set(keywords[:5]))